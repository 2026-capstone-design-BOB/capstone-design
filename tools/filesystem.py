"""
파일시스템 도구
파일/폴더 생성, 파일 탐색, 최근 파일 열기
"""

import os
import re
import glob
import subprocess
from datetime import datetime
from langchain_core.tools import tool


# 지원하는 위치 매핑
LOCATION_MAP: dict[str, str] = {
    "desktop":   os.path.join(os.path.expanduser("~"), "Desktop"),
    "바탕화면":   os.path.join(os.path.expanduser("~"), "Desktop"),
    "downloads": os.path.join(os.path.expanduser("~"), "Downloads"),
    "다운로드":   os.path.join(os.path.expanduser("~"), "Downloads"),
    "documents": os.path.join(os.path.expanduser("~"), "Documents"),
    "문서":       os.path.join(os.path.expanduser("~"), "Documents"),
    "pictures":  os.path.join(os.path.expanduser("~"), "Pictures"),
    "사진":       os.path.join(os.path.expanduser("~"), "Pictures"),
    "home":      os.path.expanduser("~"),
    "홈":         os.path.expanduser("~"),
}


def _resolve_location(location: str) -> str | None:
    """위치 문자열을 실제 경로로 변환.
    지원 형식:
    - "바탕화면", "desktop" 등 키워드
    - "C:/..." 절대 경로 (존재 여부 무관)
    - "바탕화면/서브폴더", "desktop/subfolder" 형식
    """
    key = location.lower().strip()
    path = LOCATION_MAP.get(key)
    if path:
        return path
    # 절대 경로면 그대로 사용 (존재하지 않아도 반환 — create_file/folder가 생성)
    if os.path.isabs(location):
        return location
    # "바탕화면/서브폴더" 형식 처리
    parts = location.replace("\\", "/").split("/", 1)
    if len(parts) == 2:
        base = LOCATION_MAP.get(parts[0].lower().strip())
        if base:
            return os.path.join(base, parts[1])
    return None


@tool
def create_file(name: str, location: str = "desktop", content: str = "") -> str:
    """
    파일을 생성합니다.
    name: 파일명 (확장자 포함, 예: 메모.txt, 보고서.docx)
    location: 저장 위치 — 기본 바탕화면
      · 키워드: desktop/바탕화면, downloads/다운로드, documents/문서
      · 서브폴더: "바탕화면/폴더명" 형식 가능 (예: "바탕화면/프로젝트", "downloads/새폴더")
    content: 파일 내용 (선택, 기본 빈 파일)
    """
    base = _resolve_location(location)
    if not base:
        return f"✗ '{location}'은(는) 지원하지 않는 위치입니다. (desktop, downloads, documents 중 선택)"

    path = os.path.join(base, name)
    try:
        os.makedirs(os.path.dirname(path) or base, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        return f"✓ '{name}' 파일을 {location}에 생성했습니다.\n경로: {path}"
    except Exception as e:
        return f"✗ 파일 생성 실패: {e}"


@tool
def create_folder(name: str, location: str = "desktop") -> str:
    """
    폴더를 생성합니다.
    name: 폴더명
    location: 위치 (desktop, downloads, documents) — 기본 바탕화면
    """
    base = _resolve_location(location)
    if not base:
        return f"✗ '{location}'은(는) 지원하지 않는 위치입니다."

    path = os.path.join(base, name)
    try:
        os.makedirs(path, exist_ok=True)
        return f"✓ '{name}' 폴더를 {location}에 생성했습니다.\n경로: {path}"
    except Exception as e:
        return f"✗ 폴더 생성 실패: {e}"


@tool
def find_file(name: str = "", extension: str = "", location: str = "downloads") -> str:
    """
    파일을 탐색합니다.
    name: 파일명 또는 키워드 (선택)
    extension: 확장자 (예: pdf, txt, docx) — 선택
    location: 탐색 위치 (desktop, downloads, documents) — 기본 다운로드
    """
    base = _resolve_location(location)
    if not base:
        return f"✗ '{location}'은(는) 지원하지 않는 위치입니다."

    # 검색 패턴 구성
    if name and extension:
        pattern = f"*{name}*.{extension}"
    elif name:
        pattern = f"*{name}*"
    elif extension:
        pattern = f"*.{extension}"
    else:
        return "✗ 파일명 또는 확장자를 지정해주세요."

    # recursive=True로 한 번만 검색 (하위 폴더 포함, 중복 없음)
    matches = glob.glob(os.path.join(base, "**", pattern), recursive=True)

    if not matches:
        return f"✗ '{location}'에서 조건에 맞는 파일을 찾지 못했습니다."

    result_lines = [f"✓ {len(matches)}개 파일을 찾았습니다:"]
    for i, m in enumerate(matches[:10], 1):
        result_lines.append(f"  {i}. {os.path.basename(m)}")
    if len(matches) > 10:
        result_lines.append(f"  ... 외 {len(matches)-10}개")

    return "\n".join(result_lines)


@tool
def open_recent_file() -> str:
    """
    최근에 열었던 파일 목록을 보여주고 탐색기로 최근 파일 폴더를 엽니다.
    """
    recent_path = os.path.join(
        os.environ.get("APPDATA", ""),
        "Microsoft", "Windows", "Recent"
    )
    try:
        subprocess.Popen(["explorer", recent_path])
        return f"✓ 최근 파일 폴더를 열었습니다."
    except Exception as e:
        return f"✗ 최근 파일 열기 실패: {e}"


@tool
def open_file(file_path: str, app: str = "") -> str:
    """
    파일을 기본 앱 또는 지정한 앱으로 엽니다.
    파일 경로가 있을 때 사용합니다. 앱만 실행(파일 없이)할 때는 open_app을 사용하세요.
    "파일 만들고 메모장으로 열어줘" → create_file() 후 open_file(file_path="파일명", app="notepad")
    "todo.txt를 메모장으로 열어줘" → open_file(file_path="todo.txt", app="notepad")
    file_path: 파일 경로 또는 파일명 (바탕화면·문서·다운로드 상대경로 가능)
               예: "메모.txt", "바탕화면/보고서.docx", "C:/Users/user/Desktop/todo.txt"
    app: 열 앱 이름 (예: notepad, chrome, excel). 비워두면 기본 앱으로 열기.
    """
    # 상대 경로 해석 (바탕화면, 문서, 다운로드)
    resolved = _resolve_location_in_path(file_path)

    if not os.path.exists(resolved):
        return f"✗ 파일을 찾을 수 없습니다: {resolved}"

    try:
        if app:
            # 앱 이름으로 실행 파일 찾기
            from tools.app_control import _normalize, APP_PROCESS_MAP, _resolve_path
            app_key = _normalize(app)
            exe_list = APP_PROCESS_MAP.get(app_key, [f"{app_key}.exe"])

            # 직접 exe 이름으로 시도
            import shutil
            exe_name = exe_list[0]
            exe_path = shutil.which(exe_name) or _resolve_path(app_key)

            if exe_path:
                subprocess.Popen([exe_path, resolved])
            else:
                # fallback: shell 명령
                subprocess.Popen([exe_name, resolved], shell=True)
        else:
            os.startfile(resolved)

        return f"✓ '{os.path.basename(resolved)}' 파일을 열었습니다."
    except Exception as e:
        return f"✗ 파일 열기 실패: {e}"


def _resolve_location_in_path(file_path: str) -> str:
    """경로 문자열을 실제 경로로 변환. 상대 위치명(바탕화면 등) 처리."""
    if os.path.isabs(file_path):
        return file_path

    parts = file_path.replace("\\", "/").split("/", 1)
    if len(parts) == 2:
        base = _resolve_location(parts[0])
        if base:
            return os.path.join(base, parts[1])

    for loc in ["바탕화면", "문서", "다운로드"]:
        base = _resolve_location(loc)
        if base:
            candidate = os.path.join(base, file_path)
            if os.path.exists(candidate):
                return candidate

    return file_path


@tool
def write_excel(filename: str, headers: str, rows: str, location: str = "desktop") -> str:
    """
    엑셀(.xlsx) 파일을 생성합니다. 검색·비교 데이터를 표 형식으로 저장할 때 사용합니다.
    filename: 파일명 (예: 에어컨비교.xlsx)
    headers: 열 제목, 쉼표로 구분 (예: "모델명,가격,용량,에너지효율등급")
    rows: 행 데이터. 행은 줄바꿈(\\n)으로, 열은 쉼표로 구분.
          (예: "삼성 무풍 에어컨,1200000원,18평형,1등급\\nLG 휘센,1100000원,16평형,1등급")
    location: 저장 위치 (desktop/바탕화면, downloads/다운로드, documents/문서)
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return "✗ write_excel 사용을 위해 openpyxl이 필요합니다. (pip install openpyxl)"

    base = _resolve_location(location)
    if not base:
        return f"✗ '{location}'은(는) 지원하지 않는 위치입니다."

    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    path = os.path.join(base, filename)

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "데이터"

        # 헤더 행
        header_list = [h.strip() for h in headers.split(",")]
        ws.append(header_list)

        # 헤더 스타일 (파란 배경 + 흰 글씨)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # 데이터 행
        for line in rows.strip().splitlines():
            if line.strip():
                ws.append([col.strip() for col in line.split(",")])

        # 열 너비 자동 조정
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        os.makedirs(base, exist_ok=True)
        wb.save(path)
        return f"✓ '{filename}' 엑셀 파일을 저장했습니다.\n경로: {path}"

    except Exception as e:
        return f"✗ 엑셀 파일 생성 실패: {e}"


# ── 위험 동작: 삭제 (HITL 승인 대상) ──────────────────────────────
# 이 도구들은 절대 캐시/라우터로 처리하지 않고, 그래프 hitl 노드에서
# 사용자 승인을 받은 뒤에만 실행된다 (P2).

_PROTECTED_SUBSTR = ("\\windows", "/windows", "system32", "syswow64",
                     "\\program files", "/program files")


_DRIVE_ROOT_RE = re.compile(r'^[a-zA-Z]:[\\/]?$')   # C:  C:\  C:/


def _is_protected_path(path: str) -> bool:
    """시스템/보호 경로 또는 드라이브 루트면 True (삭제 금지)."""
    raw = (path or "").strip()
    if _DRIVE_ROOT_RE.match(raw):       # 드라이브 루트 (플랫폼 무관)
        return True
    p = os.path.abspath(path)
    drive, tail = os.path.splitdrive(p)
    if tail in ("", "\\", "/"):         # 드라이브 루트(Windows)
        return True
    low = p.lower()
    return any(s in low for s in _PROTECTED_SUBSTR)


def _to_trash(path: str) -> str:
    """가능하면 휴지통으로 이동, 실패 시 영구 삭제. 방식('trash'|'permanent') 반환."""
    try:
        from send2trash import send2trash as _s2t
        _s2t(path)
        return "trash"
    except Exception:
        if os.path.isdir(path):
            import shutil
            shutil.rmtree(path)
        else:
            os.remove(path)
        return "permanent"


@tool
def delete_file(file_path: str) -> str:
    """파일을 삭제합니다(가능하면 휴지통으로 이동). 되돌리기 어려운 위험 동작이라
    반드시 사용자 승인을 받은 뒤 실행됩니다.
    file_path: 파일 경로 또는 상대경로(바탕화면/문서/다운로드).
    """
    resolved = _resolve_location_in_path(file_path)
    if not os.path.exists(resolved):
        return f"✗ 파일을 찾을 수 없습니다: {resolved}"
    if os.path.isdir(resolved):
        return f"✗ '{os.path.basename(resolved)}'은(는) 폴더예요. 폴더는 delete_folder를 쓰세요."
    if _is_protected_path(resolved):
        return "✗ 시스템/보호 경로의 파일은 삭제할 수 없어요."
    try:
        how = _to_trash(resolved)
        where = "휴지통으로 옮겼어요" if how == "trash" else "삭제했어요"
        return f"✓ '{os.path.basename(resolved)}' {where}."
    except Exception as e:
        return f"✗ 삭제 실패: {e}"


@tool
def delete_folder(folder_path: str) -> str:
    """폴더를 삭제합니다(가능하면 휴지통으로 이동). 되돌리기 어려운 위험 동작이라
    반드시 사용자 승인을 받은 뒤 실행됩니다.
    folder_path: 폴더 경로 또는 상대경로(바탕화면/문서/다운로드).
    """
    resolved = _resolve_location_in_path(folder_path)
    if not os.path.exists(resolved):
        return f"✗ 폴더를 찾을 수 없습니다: {resolved}"
    if not os.path.isdir(resolved):
        return f"✗ '{os.path.basename(resolved)}'은(는) 파일이에요. delete_file을 쓰세요."
    if _is_protected_path(resolved):
        return "✗ 시스템/보호 경로의 폴더는 삭제할 수 없어요."
    try:
        how = _to_trash(resolved)
        where = "휴지통으로 옮겼어요" if how == "trash" else "삭제했어요"
        return f"✓ '{os.path.basename(resolved)}' 폴더를 {where}."
    except Exception as e:
        return f"✗ 삭제 실패: {e}"
